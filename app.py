#!/usr/bin/env python3
"""
PRISM Risk Assessment Tool — Flask/SQLite backend

Quick start:
    pip install -r requirements.txt
    python manage.py init-db               # create tables (first run only)
    python manage.py add-user <username>   # create first analyst account
    gunicorn -w 2 -b 127.0.0.1:5000 app:app

Changelog:
    - Added PUT /api/assessments/<id> so the frontend can update an
      in-progress assessment instead of inserting a new row on every
      score change.  The frontend should POST once to create a draft,
      store the returned id, then PUT on each subsequent change and on
      final submission.  authenticated_as and created_at are never
      overwritten by a PUT — only analyst-supplied fields change.
    - Fixed CSV export: level columns (human-readable tier labels per
      CARVER dimension) are now included; pre-assessment question headers
      are renamed to match the actual form questions; CARVER column
      headers clarified with full dimension names.
"""
import csv
import importlib.metadata
import io
import os
import re
import secrets
import sqlite3
from datetime import datetime, timezone
from functools import wraps

from flask import Flask, Response, g, jsonify, render_template, request
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash
import nh3

app = Flask(__name__)

# Trust one hop of X-Forwarded-For/Proto so the rate-limiter and HSTS logic
# see the real client IP when running behind nginx.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

# Load a stable secret key from the environment (set PRISM_SECRET_KEY in production).
# Falls back to a random key so gunicorn workers each get a unique value — fine for
# current usage (no cross-worker sessions), but set the env var for production.
app.config["SECRET_KEY"] = os.environ.get("PRISM_SECRET_KEY") or secrets.token_hex(32)

# 60 req/min per IP.  With multiple gunicorn workers the effective ceiling is
# workers × 60 because each worker tracks its own in-memory counters; switch
# storage_uri to "redis://localhost:6379" for a shared limit across workers.
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=["60 per minute"],
    storage_uri="memory://",
)

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "carver.db")

# ── Database ──────────────────────────────────────────────────────────────────

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        # WAL mode allows concurrent reads while a write is in progress —
        # important for a shared multi-analyst tool.
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db


@app.teardown_appcontext
def close_db(_):
    db = g.pop("db", None)
    if db:
        db.close()


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            username   TEXT    NOT NULL UNIQUE COLLATE NOCASE,
            password   TEXT    NOT NULL,
            created_at TEXT    DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS assessments (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            -- Server-stamped; cannot be altered by the client.
            created_at       TEXT    DEFAULT (datetime('now')),
            -- The authenticated login that submitted this record.
            authenticated_as TEXT    NOT NULL,
            -- Analyst-supplied fields from the form.
            date             TEXT    NOT NULL,
            analyst          TEXT    NOT NULL,
            cve              TEXT,
            vuln_name        TEXT    NOT NULL,
            system           TEXT,
            owner            TEXT,
            business_unit    TEXT,
            threat_actor     TEXT,
            mitre            TEXT,
            vpr_score        TEXT,
            notes            TEXT,
            -- Pre-assessment yes/no answers.
            -- pre_q1 = "Is this vulnerability being actively exploited in the wild?"
            -- pre_q2 = "Is a public proof-of-concept (PoC) exploit available?"
            -- pre_q3 = "Is the affected asset externally accessible / internet-facing?"
            pre_q1           TEXT    NOT NULL,
            pre_q2           TEXT    NOT NULL,
            pre_q3           TEXT    NOT NULL,
            -- Individual CARVER scores (1–5 each).
            score_c          INTEGER NOT NULL,  -- Criticality
            score_a          INTEGER NOT NULL,  -- Accessibility
            score_r1         INTEGER NOT NULL,  -- Recuperability (Recovery)
            score_v          INTEGER NOT NULL,  -- Vulnerability
            score_e          INTEGER NOT NULL,  -- Effect
            score_r2         INTEGER NOT NULL,  -- Recognizability
            -- Human-readable level labels for each dimension.
            level_c          TEXT,
            level_a          TEXT,
            level_r1         TEXT,
            level_v          TEXT,
            level_e          TEXT,
            level_r2         TEXT,
            -- Computed totals.
            total_score      INTEGER NOT NULL,
            risk_tier        TEXT    NOT NULL,
            -- Full email-safe HTML artifact stored for examiner export.
            email_html       TEXT,
            -- Soft-delete fields; NULL means the record is active.
            deleted_at       TEXT,
            deleted_by       TEXT
        );

        CREATE TABLE IF NOT EXISTS audit_log (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp     TEXT    NOT NULL DEFAULT (datetime('now')),
            actor         TEXT    NOT NULL,
            action        TEXT    NOT NULL,
            assessment_id INTEGER,
            detail        TEXT
        );
    """)
    conn.commit()

    # Column migrations — safe to re-run; silently skipped if already present.
    for stmt in (
        "ALTER TABLE assessments ADD COLUMN deleted_at TEXT",
        "ALTER TABLE assessments ADD COLUMN deleted_by TEXT",
        "ALTER TABLE assessments ADD COLUMN not_affected INTEGER DEFAULT 0",
        "ALTER TABLE assessments ADD COLUMN confirmed_by TEXT",
        "ALTER TABLE assessments ADD COLUMN note_c  TEXT",
        "ALTER TABLE assessments ADD COLUMN note_a  TEXT",
        "ALTER TABLE assessments ADD COLUMN note_r1 TEXT",
        "ALTER TABLE assessments ADD COLUMN note_v  TEXT",
        "ALTER TABLE assessments ADD COLUMN note_e  TEXT",
        "ALTER TABLE assessments ADD COLUMN note_r2 TEXT",
    ):
        try:
            conn.execute(stmt)
        except sqlite3.OperationalError:
            pass
    conn.commit()
    conn.close()


def _log_action(action, assessment_id=None, detail=None):
    """Append one row to audit_log. Caller is responsible for db.commit()."""
    db = get_db()
    db.execute(
        "INSERT INTO audit_log (actor, action, assessment_id, detail) VALUES (?, ?, ?, ?)",
        (g.current_user, action, assessment_id, detail),
    )


# ── HTTP Basic Auth ───────────────────────────────────────────────────────────

REALM = "PRISM Risk Assessment Tool"


def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not auth.username or not auth.password:
            return Response(
                "Authentication required.",
                401,
                {"WWW-Authenticate": f'Basic realm="{REALM}"'},
            )
        db = get_db()
        row = db.execute(
            "SELECT password FROM users WHERE username = ?", (auth.username,)
        ).fetchone()
        if not row or not check_password_hash(row["password"], auth.password):
            return Response(
                "Invalid credentials.",
                401,
                {"WWW-Authenticate": f'Basic realm="{REALM}"'},
            )
        g.current_user = auth.username

        # CSRF mitigation for state-mutating requests: require the custom header
        # that same-origin XHR sends but a cross-origin form/fetch cannot set
        # without CORS pre-flight (which we don't allow).
        if request.method in ("POST", "PUT", "DELETE"):
            if request.headers.get("X-Requested-With") != "XMLHttpRequest":
                return jsonify({"error": "Missing X-Requested-With header"}), 403

        return f(*args, **kwargs)
    return decorated


# ── Shared helpers ────────────────────────────────────────────────────────────

def _assessment_fields(data, scores, total, tier, email_html):
    """
    Return the tuple of analyst-supplied field values in the order expected by
    both INSERT and UPDATE statements.

    `scores`, `total`, `tier`, and `email_html` are passed explicitly — they
    are server-computed or server-sanitized and must never be taken directly
    from the client payload.
    """
    levels    = data.get("levels", {})
    dim_notes = data.get("dim_notes", {})
    return (
        data.get("date"),
        data.get("analyst"),
        data.get("cve") or None,
        data.get("vuln_name"),
        data.get("system") or None,
        data.get("owner") or None,
        data.get("business_unit") or None,
        data.get("threat_actor") or None,
        data.get("mitre") or None,
        data.get("vpr_score") or None,
        data.get("notes") or None,
        data.get("pre_q1"),
        data.get("pre_q2"),
        data.get("pre_q3"),
        scores["C"],
        scores["A"],
        scores["R1"],
        scores["V"],
        scores["E"],
        scores["R2"],
        levels.get("C"),
        levels.get("A"),
        levels.get("R1"),
        levels.get("V"),
        levels.get("E"),
        levels.get("R2"),
        dim_notes.get("C") or None,
        dim_notes.get("A") or None,
        dim_notes.get("R1") or None,
        dim_notes.get("V") or None,
        dim_notes.get("E") or None,
        dim_notes.get("R2") or None,
        total,
        tier,
        email_html,
        1 if data.get("not_affected") else 0,
        data.get("confirmed_by") or None,
    )


# ── Score validation ──────────────────────────────────────────────────────────

_DIMS = ("C", "A", "R1", "V", "E", "R2")


def _validate_scores(data):
    """
    Validate all six CARVER dimension scores from the request payload.
    Returns (scores_dict, total, tier) with server-computed totals.
    Raises ValueError if any score is missing or outside the 1–5 range.
    """
    raw = data.get("scores", {})
    scores = {}
    for dim in _DIMS:
        v = raw.get(dim)
        if not isinstance(v, int) or not (1 <= v <= 5):
            raise ValueError(
                f"Score '{dim}' must be an integer between 1 and 5, got {v!r}"
            )
        scores[dim] = v
    total = sum(scores.values())
    tier = "LOW" if total <= 12 else "MEDIUM" if total <= 24 else "EMERGENCY"
    return scores, total, tier


# ── HTML sanitisation ─────────────────────────────────────────────────────────

_REPORT_TAGS = frozenset({
    "html", "head", "body", "title", "meta",
    "table", "thead", "tbody", "tfoot", "tr", "td", "th",
    "div", "span", "p", "strong", "br", "a",
    "h1", "h2", "h3", "h4",
})

# Allow style on every element (email HTML is entirely inline-styled).
# All other tag-specific attributes are enumerated explicitly so no
# event handler (onclick, onload, …) can slip through.
_REPORT_ATTRS = {
    "*":     {"style", "lang"},
    "table": {"border", "cellpadding", "cellspacing", "align", "width", "bgcolor"},
    "td":    {"colspan", "width", "align", "valign", "bgcolor"},
    "th":    {"colspan", "width", "align", "valign", "bgcolor"},
    "tr":    {"bgcolor"},
    "a":     {"href"},
    "meta":  {"charset", "name", "content"},
}


def _sanitize_html(html):
    """Strip scripts, event handlers, and non-whitelisted elements from report HTML."""
    if not html:
        return None
    return nh3.clean(
        html,
        tags=_REPORT_TAGS,
        attributes=_REPORT_ATTRS,
        strip_comments=True,
    )


# ── Security headers ──────────────────────────────────────────────────────────

@app.after_request
def add_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    # Allow inline styles/scripts (required by the single-page UI) while
    # blocking external resources. setdefault lets individual routes override
    # this with a stricter policy (e.g. the report download endpoint).
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'none'; "
        "script-src 'unsafe-inline'; "
        "style-src 'unsafe-inline'; "
        "connect-src 'self'; "
        "img-src 'self' data:; "
        "frame-src 'self';",
    )
    # HSTS is only meaningful over HTTPS — set it only when the request
    # arrived via a TLS-terminating proxy (nginx) or a direct TLS socket.
    if request.headers.get("X-Forwarded-Proto") == "https" or request.is_secure:
        response.headers["Strict-Transport-Security"] = (
            "max-age=31536000; includeSubDomains"
        )
    return response


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
@require_auth
def index():
    return render_template("index.html")


@app.route("/api/assessments", methods=["POST"])
@require_auth
def save_assessment():
    """
    Create a new assessment record.

    The frontend should call this exactly once — when the analyst opens a
    new assessment — and store the returned id.  All subsequent score /
    field changes should use PUT /api/assessments/<id> so that the log
    always reflects the analyst's *final* intent rather than every
    intermediate slider position.
    """
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "No JSON body provided"}), 400

    always_required = ("date", "analyst", "vuln_name")
    missing = [f for f in always_required if not data.get(f)]
    if missing:
        return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400

    not_affected = bool(data.get("not_affected"))
    if not_affected:
        if not data.get("confirmed_by"):
            return jsonify({"error": "Missing required field: confirmed_by"}), 400
        data.setdefault("pre_q1", "n/a")
        data.setdefault("pre_q2", "n/a")
        data.setdefault("pre_q3", "n/a")
        scores = {"C": 0, "A": 0, "R1": 0, "V": 0, "E": 0, "R2": 0}
        total, tier = 0, "NOT_AFFECTED"
    else:
        extra_required = ("pre_q1", "pre_q2", "pre_q3")
        missing = [f for f in extra_required if not data.get(f)]
        if missing:
            return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400
        try:
            scores, total, tier = _validate_scores(data)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    email_html = _sanitize_html(data.get("email_html"))

    db = get_db()
    cur = db.execute(
        """
        INSERT INTO assessments (
            authenticated_as,
            date, analyst, cve, vuln_name, system, owner,
            business_unit, threat_actor, mitre, vpr_score, notes,
            pre_q1, pre_q2, pre_q3,
            score_c, score_a, score_r1, score_v, score_e, score_r2,
            level_c, level_a, level_r1, level_v, level_e, level_r2,
            note_c, note_a, note_r1, note_v, note_e, note_r2,
            total_score, risk_tier, email_html,
            not_affected, confirmed_by
        ) VALUES (
            ?,
            ?,?,?,?,?,?,?,?,?,?,?,
            ?,?,?,
            ?,?,?,?,?,?,
            ?,?,?,?,?,?,
            ?,?,?,?,?,?,
            ?,?,?,
            ?,?
        )
        """,
        (g.current_user, *_assessment_fields(data, scores, total, tier, email_html)),
    )
    _log_action("create", cur.lastrowid, data.get("vuln_name"))
    db.commit()
    return jsonify({"id": cur.lastrowid}), 201


@app.route("/api/assessments/<int:aid>", methods=["PUT"])
@require_auth
def update_assessment(aid):
    """
    Overwrite all analyst-supplied fields on an existing assessment.

    authenticated_as and created_at are intentionally excluded — they record
    who first created the record and when, and must not be overwritten.

    Typical flow
    ─────────────
    1. Analyst opens a new form → frontend POSTs a draft, receives id.
    2. Analyst adjusts scores / fields → frontend PUTs to /api/assessments/<id>
       on each meaningful change (e.g. on blur, or on explicit Save click).
    3. Analyst clicks Submit / Finalise → frontend PUTs one last time with the
       complete payload including email_html.

    This ensures the log row always reflects the analyst's final intent and
    the risk_tier is never stale.
    """
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "No JSON body provided"}), 400

    always_required = ("date", "analyst", "vuln_name")
    missing = [f for f in always_required if not data.get(f)]
    if missing:
        return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400

    not_affected = bool(data.get("not_affected"))
    if not_affected:
        if not data.get("confirmed_by"):
            return jsonify({"error": "Missing required field: confirmed_by"}), 400
        data.setdefault("pre_q1", "n/a")
        data.setdefault("pre_q2", "n/a")
        data.setdefault("pre_q3", "n/a")
        scores = {"C": 0, "A": 0, "R1": 0, "V": 0, "E": 0, "R2": 0}
        total, tier = 0, "NOT_AFFECTED"
    else:
        extra_required = ("pre_q1", "pre_q2", "pre_q3")
        missing = [f for f in extra_required if not data.get(f)]
        if missing:
            return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400
        try:
            scores, total, tier = _validate_scores(data)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    email_html = _sanitize_html(data.get("email_html"))

    db = get_db()

    existing = db.execute(
        "SELECT authenticated_as FROM assessments WHERE id = ? AND deleted_at IS NULL",
        (aid,),
    ).fetchone()
    if not existing:
        return jsonify({"error": "Assessment not found"}), 404
    if existing["authenticated_as"] != g.current_user:
        return jsonify({"error": "You may only edit your own assessments"}), 403

    db.execute(
        """
        UPDATE assessments SET
            date          = ?,
            analyst       = ?,
            cve           = ?,
            vuln_name     = ?,
            system        = ?,
            owner         = ?,
            business_unit = ?,
            threat_actor  = ?,
            mitre         = ?,
            vpr_score     = ?,
            notes         = ?,
            pre_q1        = ?,
            pre_q2        = ?,
            pre_q3        = ?,
            score_c       = ?,
            score_a       = ?,
            score_r1      = ?,
            score_v       = ?,
            score_e       = ?,
            score_r2      = ?,
            level_c       = ?,
            level_a       = ?,
            level_r1      = ?,
            level_v       = ?,
            level_e       = ?,
            level_r2      = ?,
            note_c        = ?,
            note_a        = ?,
            note_r1       = ?,
            note_v        = ?,
            note_e        = ?,
            note_r2       = ?,
            total_score   = ?,
            risk_tier     = ?,
            email_html    = ?,
            not_affected  = ?,
            confirmed_by  = ?
        WHERE id = ?
        """,
        (*_assessment_fields(data, scores, total, tier, email_html), aid),
    )
    _log_action("update", aid, data.get("vuln_name"))
    db.commit()
    return jsonify({"id": aid}), 200


@app.route("/api/assessments", methods=["GET"])
@require_auth
def list_assessments():
    db = get_db()
    rows = db.execute(
        """
        SELECT id, created_at, authenticated_as, date, analyst,
               cve, vuln_name, system, total_score, risk_tier
        FROM assessments
        WHERE deleted_at IS NULL
        ORDER BY created_at DESC
        """
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/assessments/<int:aid>", methods=["DELETE"])
@require_auth
def delete_assessment(aid):
    db = get_db()
    row = db.execute(
        "SELECT authenticated_as, vuln_name FROM assessments WHERE id = ? AND deleted_at IS NULL",
        (aid,),
    ).fetchone()
    if not row:
        return jsonify({"error": "Assessment not found"}), 404
    if row["authenticated_as"] != g.current_user:
        return jsonify({"error": "You may only delete your own assessments"}), 403
    db.execute(
        "UPDATE assessments SET deleted_at = datetime('now'), deleted_by = ? WHERE id = ?",
        (g.current_user, aid),
    )
    _log_action("delete", aid, row["vuln_name"])
    db.commit()
    return "", 204


@app.route("/api/assessments/export")
@require_auth
def export_csv():
    """
    Export assessments as CSV.

    Pass ?today=1 to restrict the export to assessments dated today
    (mirrors the same flag on the .xlsx endpoint).

    Column order and headers are documented below.  CARVER dimension scores
    and their human-readable level labels are exported side-by-side so the
    CSV is self-contained without needing the web UI to interpret scores.

    Pre-assessment question headers match the exact question text shown in
    the form so that the CSV is unambiguous without needing to cross-reference
    the application:
        pre_q1 → "Actively Exploited in the Wild? (yes/no)"
        pre_q2 → "Public PoC Exploit Available? (yes/no)"
        pre_q3 → "Asset Externally Accessible / Internet-Facing? (yes/no)"
    """
    today_only = request.args.get("today") == "1"
    today_date = datetime.now().strftime("%Y-%m-%d")

    # CAST(? AS INTEGER) = 0  →  today_only is False  →  no date filter (all rows)
    # CAST(? AS INTEGER) = 1  →  today_only is True   →  restrict to date = today_date
    db = get_db()
    rows = db.execute(
        """
        SELECT
            id, created_at, authenticated_as, date, analyst,
            cve, vuln_name, system, not_affected, confirmed_by, owner, business_unit,
            threat_actor, mitre, vpr_score, notes,
            -- Pre-assessment flags (in form order)
            pre_q1, pre_q2, pre_q3,
            -- CARVER scores and their human-readable level labels,
            -- exported as adjacent pairs so the CSV is self-documenting.
            score_c,  level_c,  note_c,
            score_a,  level_a,  note_a,
            score_r1, level_r1, note_r1,
            score_v,  level_v,  note_v,
            score_e,  level_e,  note_e,
            score_r2, level_r2, note_r2,
            total_score, risk_tier
        FROM assessments
        WHERE deleted_at IS NULL
          AND (CAST(? AS INTEGER) = 0 OR date = ?)
        ORDER BY created_at DESC
        """,
        (int(today_only), today_date),
    ).fetchall()

    if today_only and not rows:
        return jsonify({"error": f"No assessments recorded for {today_date}."}), 404

    out = io.StringIO()
    w = csv.writer(out)

    # Headers must mirror the SELECT column order exactly.
    # Pre-assessment question text is spelled out in full so the CSV is
    # self-explanatory without the web UI.
    w.writerow([
        "ID",
        "Server Timestamp (UTC)",
        "Authenticated As",
        "Assessment Date",
        "Analyst",
        "CVE",
        "Vulnerability / Threat",
        "Affected System",
        "Not Affected",
        "Confirmed By",
        "Asset Owner",
        "Business Unit",
        "Threat Actor",
        "MITRE ATT&CK",
        "VPR Score",
        "Source Reports",
        # Pre-assessment questions — full text matches the form labels.
        "Actively Exploited in the Wild? (yes/no)",
        "Public PoC Exploit Available? (yes/no)",
        "Asset Externally Accessible / Internet-Facing? (yes/no)",
        # CARVER dimensions: score then level label, six pairs.
        "C – Criticality (Score)",        "C – Criticality (Level)",        "C – Criticality (Notes)",
        "A – Accessibility (Score)",      "A – Accessibility (Level)",      "A – Accessibility (Notes)",
        "R – Recuperability (Score)",     "R – Recuperability (Level)",     "R – Recuperability (Notes)",
        "V – Vulnerability (Score)",      "V – Vulnerability (Level)",      "V – Vulnerability (Notes)",
        "E – Effect (Score)",             "E – Effect (Level)",             "E – Effect (Notes)",
        "R – Recognizability (Score)",    "R – Recognizability (Level)",    "R – Recognizability (Notes)",
        # Totals.
        "Total Score (/30)",
        "Risk Tier",
    ])

    for r in rows:
        w.writerow([
            r["id"],
            r["created_at"],
            r["authenticated_as"],
            r["date"],
            r["analyst"],
            r["cve"],
            r["vuln_name"],
            r["system"],
            "Yes" if r["not_affected"] else "No",
            r["confirmed_by"] or "",
            r["owner"],
            r["business_unit"],
            r["threat_actor"],
            r["mitre"],
            r["vpr_score"],
            r["notes"] or "",
            # Pre-assessment answers (order matches pre_q1/2/3 above).
            r["pre_q1"],
            r["pre_q2"],
            r["pre_q3"],
            # CARVER score+level pairs.
            r["score_c"],  r["level_c"],  r["note_c"]  or "",
            r["score_a"],  r["level_a"],  r["note_a"]  or "",
            r["score_r1"], r["level_r1"], r["note_r1"] or "",
            r["score_v"],  r["level_v"],  r["note_v"]  or "",
            r["score_e"],  r["level_e"],  r["note_e"]  or "",
            r["score_r2"], r["level_r2"], r["note_r2"] or "",
            r["total_score"],
            r["risk_tier"],
        ])

    filename = (
        f"prism_assessments_{today_date}.csv"
        if today_only
        else f"prism_assessments_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"
    )
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/api/assessments/export/xlsx")
@require_auth
def export_xlsx():
    """
    Export all assessments as a formatted Excel workbook (.xlsx).

    Designed for SharePoint's Create List from Excel / Import Spreadsheet
    workflow.  The workbook includes:
      - Frozen header row (navy / white, matching the app colour scheme)
      - Risk Tier cells colour-coded by severity
      - Auto-fitted column widths (capped at 55 characters)
      - Score columns stored as integers so SharePoint treats them as numbers
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return jsonify({"error": "openpyxl is not installed in this environment"}), 500

    today_only = request.args.get("today") == "1"
    today_date = datetime.now().strftime("%Y-%m-%d")

    # CAST(? AS INTEGER) = 0  →  today_only is False  →  no date filter (all rows)
    # CAST(? AS INTEGER) = 1  →  today_only is True   →  restrict to date = today_date
    db   = get_db()
    rows = db.execute(
        """
        SELECT
            id, created_at, authenticated_as, date, analyst,
            cve, vuln_name, system, not_affected, confirmed_by, owner, business_unit,
            threat_actor, mitre, vpr_score, notes,
            pre_q1, pre_q2, pre_q3,
            score_c,  level_c,  note_c,
            score_a,  level_a,  note_a,
            score_r1, level_r1, note_r1,
            score_v,  level_v,  note_v,
            score_e,  level_e,  note_e,
            score_r2, level_r2, note_r2,
            total_score, risk_tier
        FROM assessments
        WHERE deleted_at IS NULL
          AND (CAST(? AS INTEGER) = 0 OR date = ?)
        ORDER BY created_at DESC
        """,
        (int(today_only), today_date),
    ).fetchall()

    if today_only and not rows:
        return jsonify({"error": f"No assessments recorded for {today_date}."}), 404

    headers = [
        "ID", "Server Timestamp (UTC)", "Authenticated As",
        "Assessment Date", "Analyst",
        "CVE", "Vulnerability / Threat", "Affected System",
        "Not Affected", "Confirmed By",
        "Asset Owner", "Business Unit", "Threat Actor",
        "MITRE ATT&CK", "VPR Score",
        "Source Reports",
        "Actively Exploited in the Wild?",
        "Public PoC Exploit Available?",
        "Asset Externally Accessible / Internet-Facing?",
        "C – Criticality (Score)",      "C – Criticality (Level)",      "C – Criticality (Notes)",
        "A – Accessibility (Score)",    "A – Accessibility (Level)",    "A – Accessibility (Notes)",
        "R – Recuperability (Score)",   "R – Recuperability (Level)",   "R – Recuperability (Notes)",
        "V – Vulnerability (Score)",    "V – Vulnerability (Level)",    "V – Vulnerability (Notes)",
        "E – Effect (Score)",           "E – Effect (Level)",           "E – Effect (Notes)",
        "R – Recognizability (Score)",  "R – Recognizability (Level)",  "R – Recognizability (Notes)",
        "Total Score (/30)", "Risk Tier",
    ]

    tier_fills = {
        "LOW":          PatternFill("solid", fgColor="D1FAE5"),
        "MEDIUM":       PatternFill("solid", fgColor="FEF3C7"),
        "EMERGENCY":    PatternFill("solid", fgColor="FEE2E2"),
        "NOT_AFFECTED": PatternFill("solid", fgColor="DBEAFE"),
    }
    tier_col = headers.index("Risk Tier") + 1  # openpyxl columns are 1-based

    wb = Workbook()
    ws = wb.active
    ws.title = "PRISM Assessments"

    # Header row — navy background, white bold text.
    ws.append(headers)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1A2744")
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=False)

    ws.freeze_panes = "A2"  # keep headers visible while scrolling

    # Data rows.
    for r in rows:
        ws.append([
            r["id"],
            r["created_at"],
            r["authenticated_as"],
            r["date"],
            r["analyst"],
            r["cve"]          or "",
            r["vuln_name"],
            r["system"]       or "",
            "Yes" if r["not_affected"] else "No",
            r["confirmed_by"] or "",
            r["owner"]        or "",
            r["business_unit"] or "",
            r["threat_actor"] or "",
            r["mitre"]        or "",
            r["vpr_score"]    or "",
            r["notes"]        or "",
            r["pre_q1"],
            r["pre_q2"],
            r["pre_q3"],
            r["score_c"],  r["level_c"]  or "", r["note_c"]  or "",
            r["score_a"],  r["level_a"]  or "", r["note_a"]  or "",
            r["score_r1"], r["level_r1"] or "", r["note_r1"] or "",
            r["score_v"],  r["level_v"]  or "", r["note_v"]  or "",
            r["score_e"],  r["level_e"]  or "", r["note_e"]  or "",
            r["score_r2"], r["level_r2"] or "", r["note_r2"] or "",
            r["total_score"],
            r["risk_tier"],
        ])
        tier = r["risk_tier"]
        if tier in tier_fills:
            ws.cell(row=ws.max_row, column=tier_col).fill = tier_fills[tier]

    # Auto-fit column widths (capped so wide notes columns don't blow out the sheet).
    for col_cells in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 55)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = (
        f"prism_assessments_{today_date}.xlsx"
        if today_only
        else f"prism_assessments_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/api/assessments/<int:aid>/report")
@require_auth
def download_report(aid):
    db = get_db()
    row = db.execute(
        "SELECT email_html, vuln_name, cve, date FROM assessments"
        " WHERE id = ? AND deleted_at IS NULL",
        (aid,),
    ).fetchone()
    if not row or not row["email_html"]:
        return jsonify({"error": "Report not found"}), 404

    def slugify(text):
        """Collapse any non-alphanumeric character (except hyphens) to an
        underscore, strip leading/trailing underscores, and lowercase."""
        return "".join(
            c if c.isalnum() or c == "-" else "_"
            for c in (text or "")
        ).strip("_").lower()

    # Build segments in order: date → CVE (optional) → vuln name.
    # Joining non-empty segments with "_" keeps the name readable and
    # ensures every saved file is unique by date + identifier.
    segments = [
        row["date"] or "undated",
        slugify(row["cve"]) if row["cve"] else "",
        slugify(row["vuln_name"]) or "report",
    ]
    filename = "prism_" + "_".join(s for s in segments if s) + ".html"

    return Response(
        row["email_html"],
        mimetype="text/html",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            # Scripts are meaningless in an email report; block them outright.
            "Content-Security-Policy": (
                "default-src 'none'; style-src 'unsafe-inline'; img-src data: cid:;"
            ),
        },
    )


# ── Current user ─────────────────────────────────────────────────────────────

@app.route("/api/me")
@require_auth
def me():
    """Return the authenticated username so the frontend can display it."""
    return jsonify({"username": g.current_user})


# ── SBOM ──────────────────────────────────────────────────────────────────────

@app.route("/api/sbom")
@require_auth
def sbom():
    """Return direct dependencies with resolved installed versions (CISA SBOM practice)."""
    req_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "requirements.txt")
    components = []
    try:
        with open(req_path) as fh:
            for raw in fh:
                line = raw.strip()
                if not line or line.startswith("#"):
                    continue
                name = re.split(r"[\s><=!~\[;]", line)[0]
                try:
                    version = importlib.metadata.version(name)
                except importlib.metadata.PackageNotFoundError:
                    version = "not installed"
                components.append({
                    "name":       name,
                    "version":    version,
                    "constraint": line,
                })
    except FileNotFoundError:
        pass
    return jsonify({
        "generated":  datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "components": components,
    })


# ── Logout ────────────────────────────────────────────────────────────────────

@app.route("/logout")
def logout():
    """Return 401 to flush the browser's cached Basic Auth credentials."""
    return Response(
        '<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">'
        "<title>Signed out — PRISM</title></head>"
        '<body style="font-family:sans-serif;text-align:center;padding:4rem;">'
        "<h2>You have been signed out.</h2>"
        '<p><a href="/">Sign in again</a></p>'
        "</body></html>",
        401,
        {
            "WWW-Authenticate": f'Basic realm="{REALM}"',
            "Content-Type": "text/html; charset=utf-8",
        },
    )


# ── Startup ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("PRISM Risk Assessment Tool")
    print(f"  Database : {DB_PATH}")
    print("  Listening: http://0.0.0.0:5000")
    print()
    app.run(host="0.0.0.0", port=5000, debug=False)
